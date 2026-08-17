"""Excel extraction: openpyxl for structure/comments, preserving merged cells."""
import openpyxl
from loguru import logger

from models.financial import CellComment, ExtractedWorkbook, SheetData


# Content-fallback phrases, deliberately multi-word: a bare word like
# "cash"/"balance"/"profit" shows up incidentally inside any statement (a
# cash flow statement's own "bank balances" reconciliation line, a balance
# sheet's "cash and cash equivalents" row, "other equity" referencing prior
# profits) — matching those alone over a whole sheet's content misclassifies
# too easily to be safe. These phrases are specific enough to a statement's
# own structure that they don't turn up in the others.
# Checked in this order: cash_flow's phrases are the most structurally
# unique (a P&L or balance sheet never contains "cash flow from operating
# activities" as a literal heading), so they're checked before pnl's —
# an indirect-method cash flow statement reconciles from "profit before
# tax" as its own opening line, which would otherwise false-positive as pnl
# if checked first.
_CONTENT_PHRASES: dict[str, tuple[str, ...]] = {
    "cash_flow": (
        "cash flow from operating", "cash flow from investing",
        "cash flow from financing", "net increase/(decrease) in cash",
    ),
    "balance_sheet": ("balance sheet", "total equity and liabilities"),
    "pnl": ("profit and loss", "statement of profit", "revenue from operations", "total income"),
}


def _detect_sheet_type(sheet_name: str, content_sample: str = "") -> str:
    """Classify a sheet as balance_sheet/pnl/cash_flow/unknown so map_all_items
    can stop a cash-flow reconciliation line like "(Increase)/decrease in
    other financial assets" from being accepted as the balance sheet's
    actual "other financial assets" balance (see _map_rows' cross-statement
    guard) — same wording, entirely different figure.

    Sheet name is checked first (cheap, usually reliable when a workbook was
    authored directly in Excel), but real-world files are often produced by
    a generic PDF-to-Excel conversion tool that names every sheet "Table 1",
    "Table 2", etc. — meaningless for classification, even when the
    statement's own title ("Balance Sheet", "Statement of Cash Flow"...) is
    still sitting somewhere in the sheet's own content, or — as seen on a
    real user-submitted file — the title is missing from the export
    entirely and only the statement's own line items give it away. Falling
    back to content (via _CONTENT_PHRASES) is what makes the guard actually
    engage for those files instead of silently no-op'ing for the whole
    workbook."""
    name = sheet_name.lower()
    if "balance" in name:
        return "balance_sheet"
    if "profit" in name or "loss" in name or "p&l" in name or "pnl" in name:
        return "pnl"
    if "cash" in name:
        return "cash_flow"

    content = content_sample.lower()
    for stype, needles in _CONTENT_PHRASES.items():
        if any(n in content for n in needles):
            return stype
    return "unknown"


def _apply_merged_cells(ws) -> None:
    """Copy the top-left value of each merged range to every cell in the range."""
    for merged_range in list(ws.merged_cells.ranges):
        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left.value
        for row in ws.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                cell.value = value


def _extract_comments(ws, sheet_name: str) -> list[CellComment]:
    comments: list[CellComment] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                comments.append(
                    CellComment(
                        sheet=sheet_name,
                        cell_ref=cell.coordinate,
                        value=cell.value,
                        comment_text=cell.comment.text,
                        comment_author=cell.comment.author,
                    )
                )
    return comments


def extract_excel(excel_path: str) -> ExtractedWorkbook:
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        logger.error(f"Failed to load workbook {excel_path}: {e}")
        return ExtractedWorkbook(sheets=[], all_comments=[])

    sheets: list[SheetData] = []
    all_comments: list[CellComment] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            _apply_merged_cells(ws)
        except Exception as e:
            logger.warning(f"Merged-cell resolution failed for sheet {sheet_name}: {e}")

        comments = _extract_comments(ws, sheet_name)
        all_comments.extend(comments)

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            sheets.append(
                SheetData(
                    sheet_name=sheet_name,
                    sheet_type=_detect_sheet_type(sheet_name),
                    headers=[],
                    rows=[],
                    comments=comments,
                )
            )
            continue

        # Content fallback for _detect_sheet_type: a statement's own title
        # ("Balance Sheet as at...") usually appears in the first couple of
        # rows when present at all, but some exports omit it entirely and
        # start straight at the column header row — scanning the whole
        # sheet (still small, well under a hundred rows for any of these
        # statements) means a title-less P&L sheet still gets classified
        # correctly from its own line items ("Profit before tax", "Revenue
        # from Operations"), not just left unrecognized.
        content_sample = " ".join(
            str(cell) for row in all_rows for cell in row if cell is not None
        )

        headers = [str(h) if h is not None else "" for h in all_rows[0]]
        rows = [list(row) for row in all_rows[1:]]

        sheets.append(
            SheetData(
                sheet_name=sheet_name,
                sheet_type=_detect_sheet_type(sheet_name, content_sample),
                headers=headers,
                rows=rows,
                comments=comments,
            )
        )

    return ExtractedWorkbook(sheets=sheets, all_comments=all_comments)
